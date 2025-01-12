from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from .forms import Taskform
from .models import task
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.contrib.auth.decorators import login_required


# Create your views here.


def home(request):
    return render(request, "home.html")


def signup(request):
    if request.method == "GET":
        print("enviando formulario")
        return render(request, "signup.html", {"form": UserCreationForm})
    else:
        try:
            if request.POST["password1"] == request.POST["password2"]:
                user = User.objects.create_user(
                    username=request.POST["username"],
                    password=request.POST["password1"],
                )
                user.save()
                login(request, user)
                return redirect("tasks")
        except IntegrityError:
            return render(
                request,
                "signup.html",
                {"form": UserCreationForm, "error": "El usuario ya existe."},
            )
    return render(
        request,
        "signup.html",
        {"form": UserCreationForm, "error": "La contraseña no coincide."},
    )

@login_required
def tasks(request):
    tasks = task.objects.filter(user=request.user)
    return render(request, "tasks.html", {"tasks": tasks})

@login_required
def create_task(request):
    if request.method == "GET":
        return render(request, "create_task.html", {"form": Taskform()})
    else:
        try:
            form = Taskform(request.POST)
            new_task = form.save(commit=False)
            new_task.user = request.user
            new_task.save()
            print(new_task)
            return redirect("tasks")
        except ValueError:
            return render(
                request,
                "create_task.html",
                {"form": Taskform(), "error": "Por favor provee un dato valido"},
            )

@login_required
def task_detail(request, task_id):
   if request.method == 'GET':
        Task = get_object_or_404(task, pk=task_id,user=request.user)
        form = Taskform(instance=Task)
        return render(request, "task_detail.html", {"task": Task, "form": form})
   else:
       try:
           Task = get_object_or_404(task, pk=task_id, user=request.user)
           form=Taskform(request.POST, instance=Task)
           form.save()
           return redirect('tasks')
       except ValueError:
           return render(request, "task_detail.html", {"task": Task, "form": form, 'error': "Error al actualizar, intentelo de nuevo"})

@login_required           
def delete_task(request, task_id):
    Task=get_object_or_404(task,pk=task_id, user=request.user)
    if request.method =='POST':
        Task.delete()
        return redirect('tasks')

@login_required
def signout(request):
    logout(request)
    return redirect("home")


def signin(request):
    if request.method == "GET":
        return render(request, "signin.html", {"form": AuthenticationForm})
    else:
        user = authenticate(
            request,
            username=request.POST["username"],
            password=request.POST["password"],
        )
        if user is None:
            return render(
                request,
                "signin.html",
                {
                    "form": AuthenticationForm,
                    "error": "Usuario o contraseña incorrectos",
                },
            )
        else:
            login(request, user)
            return redirect("tasks")

@login_required
def download_tasks(request):
    # Obtener todas las tareas del usuario actual
    tasks = task.objects.filter(user=request.user)

    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active

    # Configurar el estilo de la fuente
    font_title = Font(name='Arial', size=18, bold=True)
    font_header = Font(name='Arial', size=18, bold=True)
    font_data = Font(name='Arial', size=18)

    # Configurar el borde
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Configurar la alineación
    alignment = Alignment(horizontal='center', vertical='center')

    # Añadir el título
    ws.append(["INVENTARIO FUNDALANAVIAL (ANZOATEGUI)"])
    ws.merge_cells('A1:D1')  # Combinar celdas para el título
    ws['A1'].font = font_title
    ws['A1'].alignment = alignment

    # Añadir la fecha
    ws.append([f"Fecha de exportación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.merge_cells('A2:D2')  # Combinar celdas para la fecha
    ws['A2'].font = font_title
    ws['A2'].alignment = alignment

    # Espacio en blanco
    ws.append([])

    # Añadir los encabezados de la tabla
    headers = ["Equipo", "Cantidad", "Serial", "Bien Nacional"]
    ws.append(headers)

    # Aplicar estilo a los encabezados
    for col in range(1, 5):  # Columnas A a D
        cell = ws.cell(row=4, column=col)
        cell.font = font_header
        cell.border = border
        cell.alignment = alignment

    # Añadir los datos de las tareas
    for t in tasks:
        ws.append([t.Equipo, t.Cantidad, t.Serial, t.Bien_Nacional])

    # Aplicar estilo a los datos
    for row in ws.iter_rows(min_row=5, max_col=4, max_row=ws.max_row):
        for cell in row:
            cell.font = font_data
            cell.border = border
            cell.alignment = alignment

    # Ajustar el tamaño de las filas a partir de la fila 3 en adelante
    for row in range(3, ws.max_row + 1):
        ws.row_dimensions[row].height = 30

    # Ajustar el ancho de las columnas A, B, C y D a 55
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 55

    # Crear la respuesta HTTP con el archivo de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventario_fundalanavial.xlsx'

    # Guardar el libro de Excel en la respuesta
    wb.save(response)

    return response